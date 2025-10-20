from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime
from dateutil import parser
from pathlib import Path
import pandas as pd


class XLSXHandler:
    _COLS = ["Actualidad Informativa", "Subido por", "Fecha","Comentarios"]
    _BASE_DIR = Path(__file__).resolve().parent
    _FILE_NAME = "news.xlsx"
    _NEWS_PATH = _BASE_DIR / "news" / _FILE_NAME

    # Build Local XLS .
    def __init__(self):
        self.df = pd.DataFrame(columns = self._COLS)
        self.df.to_excel(self._NEWS_PATH,index = False)

    # Build URL Link
    def build_url_link(self):
        self.wb = load_workbook(self._NEWS_PATH)
        self.ws = self.wb.active

        for row in self.ws.iter_rows(min_row=2, max_row= self.ws.max_row, min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="0000FF", underline="single") 
        
        self.wb.save(self._NEWS_PATH)

    # Add new to XLS
    def add_news(self,news):
        row ={
            self._COLS[0]: news["link"],
            self._COLS[1] : "CyberSecNews BOT",
            self._COLS[2] : news["published_date"],
            self._COLS[3] : news["title"]
        }

        self.df = pd.concat([self.df,pd.DataFrame([row])], ignore_index=True)
        self.df.to_excel(self._NEWS_PATH,index = False)   